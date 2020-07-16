import sys
import openpyxl
import csv
from math import sin, cos, asin, sqrt, pi
# import numpy
# import h5py

_CONST_RADIUS = 6371   # Радиус Земли(расчёт в километрах)


def execute():
    sys.stdout.write('Чтение файла...\n')
    wb = openpyxl.load_workbook('./Населенные_пункты_РФ_название_адрес_координаты1.xlsx')
    sheet = wb['Лист1']
    sheet_rows_count = sheet.max_row

    sys.stdout.write('Формирование данных...\n')

    tuple_of_places = tuple()
    array_of_places = []

    data = []
    tuple_data = tuple()

    for item in range(2, sheet.max_row + 1):
        array_of_places.append(sheet['A' + str(item)].value)
        data.append((sheet['D' + str(item)].value, sheet['E' + str(item)].value,))

    tuple_of_places = tuple(array_of_places)
    tuple_data = tuple(data)

    sheet_data_rows_count = sheet_rows_count - 1

    matrix = set_matrix_coords(tuple_data, sheet_data_rows_count)

    # Заполнение матрицы недостающими зеркальными элементами
    matrix = mirroring_update_matrix(matrix, sheet_data_rows_count)

    # Добавление горизонтального заголовка городов
    matrix = [[' ', *tuple_of_places]] + matrix

    # Добавление вертикального заголовка городов
    for x in range(1, sheet_data_rows_count):
        matrix[x] = [tuple_of_places[x - 1]] + matrix[x]

    createCSV(matrix)


# Нахождение расстояния между точками
def get_range(lat1, lon1, coordinates2):
    # Перевод широты и долготы из градусов в радианы
    lat2 = coordinates2[0] * pi / 180
    lon2 = coordinates2[1] * pi / 180

    # Вычисляем расстояние между точками по формуле Гаверсинуса на поверхности сферы.
    # Результат переводим в метры.
    sin1 = sin((lat1 - lat2) / 2)
    sin2 = sin((lon1 - lon2) / 2)
    return int((2 * _CONST_RADIUS * asin(sqrt(sin1*sin1+sin2*sin2*cos(lat1)*cos(lat2)))) * 1000)


def set_matrix_coords(tuple_data, sheet_data_rows_count):
    sys.stdout.write('Расчёт удаленности населенных пунктов...\n')
    matrix = []
    for index, data_el in enumerate(tuple_data):
        row_content = []
        lat1 = data_el[0] * pi / 180
        lon1 = data_el[1] * pi / 180
        for srn in range(index + 1, sheet_data_rows_count):
            row_content.append(get_range(lat1, lon1, tuple_data[srn]))
        matrix.append([*row_content])
        sys.stdout.write('Строк рассчитано ' + str(index + 1) + ' из ' + str(sheet_data_rows_count) + '...\r')
    sys.stdout.write('\n')
    return matrix


def mirroring_update_matrix(matrix, sheet_data_rows_count):
    # Заполнение матрицы недостающими зеркальными элементами
    for x in range(sheet_data_rows_count):
        matrix[x] = [0] * (x + 1) + matrix[x]

    for x in range(sheet_data_rows_count):
        for y in range(x + 1, sheet_data_rows_count):
            matrix[y][x] = matrix[x][y]
        sys.stdout.write('Отзеркаливание повторяющихся данных ' + str(x + 1) + ' из ' + str(sheet_data_rows_count) + '...\r')

    return matrix


def createCSV(matrix):
    sys.stdout.write('\nЗапись CSV-файла...\n')
    with open('sw_file.csv', 'w') as f:
        writer = csv.writer(f)
        for row in matrix:
            writer.writerow(row)

    sys.stdout.write('Успешно!\n')


execute()
